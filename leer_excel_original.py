"""
Script para leer y analizar el archivo Excel original
"""
import pandas as pd
import openpyxl
import os
import sys

# Ruta del archivo Excel original
excel_file = r"C:\Users\ASUS\OneDrive\Desktop\Herramienta de evaluación GPP_0 (7).xlsm"

print("=" * 70)
print("  ANÁLISIS DEL ARCHIVO EXCEL ORIGINAL")
print("=" * 70)
print()

try:
    if not os.path.exists(excel_file):
        print(f"ERROR: No se encontró el archivo")
        print(f"Ruta: {excel_file}")
        print()
        print("Por favor verifique que el archivo existe en la ubicación indicada.")
        sys.exit(1)
    
    # Leer con openpyxl para obtener información detallada
    print(f"Leyendo archivo: {os.path.basename(excel_file)}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"\nNúmero de hojas: {len(wb.sheetnames)}")
    print(f"\nHojas disponibles:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        sheet = wb[sheet_name]
        print(f"  {idx}. {sheet_name:30} ({sheet.max_row} filas x {sheet.max_column} columnas)")
    
    # Analizar hojas principales
    print("\n" + "=" * 70)
    print("  ANÁLISIS DETALLADO DE HOJAS PRINCIPALES")
    print("=" * 70)
    
    hojas_importantes = ["MATRIZ_PA", "MATRIZ _PO", "BASE GENERAL"]
    
    for sheet_name in hojas_importantes:
        if sheet_name in wb.sheetnames:
            print(f"\n{'─'*70}")
            print(f"  HOJA: {sheet_name}")
            print(f"{'─'*70}")
            
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                print(f"\nDimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
                print(f"\nPrimeras 10 filas:")
                print(df.head(10).to_string())
                
            except Exception as e:
                print(f"Error al leer con pandas: {e}")
    
    wb.close()
    
    print("\n" + "=" * 70)
    print("  ANÁLISIS COMPLETADO")
    print("=" * 70)
    print()
    print("✓ El archivo Excel ha sido analizado exitosamente")
    print()
    print("Nota: Esta aplicación web replica la funcionalidad del Excel")
    print("      de manera más intuitiva y moderna.")
    
except Exception as e:
    print(f"\nERROR al procesar el archivo: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)




