from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response
from supabase import create_client
from app.config import get_settings
from app.recomendaciones import MATRIZ_RECOMENDACIONES
from app.core.dependencies import get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

router = APIRouter(prefix="/export", tags=["export"])


def get_supabase_client():
    settings = get_settings()
    return create_client(settings.supabase_url, settings.supabase_anon_key)


def _get_score_status(pct: float) -> str:
    """Return status string based on percentage."""
    if pct >= 75:
        return "Bueno"
    elif pct >= 60:
        return "Regular"
    else:
        return "Necesita mejorar"


def _get_priority_badge(pa_pct: float, po_pct: float) -> str:
    """Return priority string based on PA/PO percentages."""
    if pa_pct < po_pct - 5:
        return "PA"
    elif po_pct < pa_pct - 5:
        return "PO"
    else:
        return "BALANCED"


def _get_recommendation_priority(rating: int) -> str:
    """Determine priority based on rating value."""
    if rating <= 2:
        return "ALTA"
    elif rating == 3:
        return "MEDIA"
    else:
        return "BAJA"


def extract_aspect_from_qid(qid: str):
    """Extract aspect name from question ID like PA_PLANEACIÓN_Análisis del contexto_0"""
    parts = qid.split("_", 2)
    if len(parts) >= 3:
        category = parts[1]
        aspect_with_index = parts[2]
        aspect = "_".join(aspect_with_index.rsplit("_", 1)[:-1])
        return (category, aspect)
    return (None, None)


def _compute_breakdown(evals: dict, matriz: dict) -> dict:
    """Compute aspect breakdown percentages from evaluation data."""
    breakdown = {}
    for aspect, categories in matriz.items():
        if aspect not in evals:
            breakdown[aspect] = 0.0
            continue
        aspect_evals = evals[aspect]
        ratings = []
        for category, questions_list in categories.items():
            for q in questions_list:
                qid = q["id"]
                if qid in aspect_evals:
                    ratings.append(aspect_evals[qid])
        if ratings:
            avg_rating = sum(ratings) / len(ratings)
            breakdown[aspect] = (avg_rating / 5) * 100
        else:
            breakdown[aspect] = 0.0
    return breakdown


def _build_questions_from_evals(evals_pa: dict, evals_po: dict) -> list:
    """Build questions array from evaluation data."""
    from app.api.matrices import MATRIZ_PA, MATRIZ_PO
    
    questions = []
    
    # Process PA evaluations
    for aspect, categories in MATRIZ_PA.items():
        if aspect not in evals_pa:
            continue
        aspect_evals = evals_pa[aspect]
        for category, questions_list in categories.items():
            for q in questions_list:
                qid = q["id"]
                if qid in aspect_evals:
                    rating = aspect_evals[qid]
                    questions.append({
                        "aspect": aspect,
                        "question": q["pregunta"],
                        "context": q.get("contexto", ""),
                        "rating": rating,
                        "percentage": (rating / 5) * 100
                    })
    
    # Process PO evaluations
    for aspect, categories in MATRIZ_PO.items():
        if aspect not in evals_po:
            continue
        aspect_evals = evals_po[aspect]
        for category, questions_list in categories.items():
            for q in questions_list:
                qid = q["id"]
                if qid in aspect_evals:
                    rating = aspect_evals[qid]
                    questions.append({
                        "aspect": aspect,
                        "question": q["pregunta"],
                        "context": q.get("contexto", ""),
                        "rating": rating,
                        "percentage": (rating / 5) * 100
                    })
    
    return questions


def _get_aspect_ratings(evals_pa: dict, evals_po: dict) -> dict:
    """Collect aspect ratings from evaluations for recommendations."""
    aspect_ratings = {}
    
    # Map aspect names to categories for recommendations lookup
    aspect_to_category = {
        "Análisis del contexto": "PLANEACIÓN",
        "Existencia de un plan estratégico": "PLANEACIÓN",
        "Estructura organizativa": "PLANEACIÓN",
        "Departamentalización": "PLANEACIÓN",
        "Procesos documentados": "PLANEACIÓN",
        "Ciclo PHVA": "PLANEACIÓN",
        "Existencia de una estructura organizativa": "ORGANIZACIÓN",
        "Procesos documentados": "ORGANIZACIÓN",
        "Liderazgo organizacional": "DIRECCIÓN",
        "Canales de comunicación efectivos": "DIRECCIÓN",
        "Cultura organizacional": "DIRECCIÓN",
        "Sistemas de control": "CONTROL",
        "Auditorías internas": "CONTROL",
        "Acciones correctivas y preventivas": "CONTROL",
        "Logística de entrada": "LOGÍSTICA DE COMPRAS",
        "Gestión de inventarios": "LOGÍSTICA DE COMPRAS",
        "Equipos": "GESTIÓN DE PRODUCCIÓN",
        "Infraestructura": "GESTIÓN DE PRODUCCIÓN",
        "Distribución de los procesos (flujo)": "GESTIÓN DE PRODUCCIÓN",
        "Planeación de la producción": "GESTIÓN DE PRODUCCIÓN",
        "Materia prima": "GESTIÓN DE PRODUCCIÓN",
        "Control de la producción": "GESTIÓN DE PRODUCCIÓN",
        "Distribución": "LOGÍSTICA EXTERNA"
    }
    
    all_evals = list(evals_pa.items()) + list(evals_po.items())
    for category, questions in all_evals:
        for qid, rating in questions.items():
            _, aspect = extract_aspect_from_qid(qid)
            if aspect and aspect in aspect_to_category:
                if aspect not in aspect_ratings:
                    aspect_ratings[aspect] = []
                aspect_ratings[aspect].append(rating)
    
    return aspect_ratings


def _create_styled_cell(ws, row, col, value, bold=False, fill_color=None, alignment=None):
    """Helper to create a styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if alignment:
        cell.alignment = alignment
    return cell


def _style_header(ws, row, num_cols):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


@router.get("/excel/{evaluation_id}")
async def export_evaluation_to_excel(evaluation_id: str, user: dict = Depends(get_current_user)):
    """Generate an Excel file with evaluation results, breakdowns, recommendations, and action plans."""
    supabase = get_supabase_client()
    
    # Fetch evaluation
    response = supabase.table("evaluations").select("*").eq("id", evaluation_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    
    evaluation = response.data[0]
    
    evals_pa = evaluation.get("evaluaciones_pa", {})
    evals_po = evaluation.get("evaluaciones_po", {})
    general_pct = evaluation.get("general_pct", 0)
    pa_pct = evaluation.get("pa_pct", 0)
    po_pct = evaluation.get("po_pct", 0)
    
    # Compute breakdowns
    from app.api.matrices import MATRIZ_PA, MATRIZ_PO
    pa_breakdown = _compute_breakdown(evals_pa, MATRIZ_PA)
    po_breakdown = _compute_breakdown(evals_po, MATRIZ_PO)
    
    # Build questions
    questions = _build_questions_from_evals(evals_pa, evals_po)
    
    # Compute priority
    priority = _get_priority_badge(pa_pct, po_pct)
    
    # Fetch recommendations
    aspect_ratings = _get_aspect_ratings(evals_pa, evals_po)
    recommendations = []
    for category, elements in MATRIZ_RECOMENDACIONES.items():
        for element, recommendation_text in elements.items():
            ratings = aspect_ratings.get(element, [])
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                rec_priority = _get_recommendation_priority(int(avg_rating))
            else:
                rec_priority = "MEDIA"
            
            recommendations.append({
                "aspect": category,
                "element": element,
                "recommendation": recommendation_text,
                "priority": rec_priority
            })
    
    # Fetch action plans
    action_plans_response = supabase.table("action_plans").select("*").eq("evaluation_id", evaluation_id).execute()
    action_plans = action_plans_response.data
    
    # Create workbook
    wb = Workbook()
    
    # ========== Sheet 1: Resumen General ==========
    ws1 = wb.active
    ws1.title = "Resumen General"
    
    # Header
    ws1.merge_cells('A1:E1')
    header_cell = ws1.cell(row=1, column=1, value="RESULTADOS DE AUDITORÍA")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Evaluation info
    eval_date = evaluation.get("fecha", "")
    if eval_date:
        if isinstance(eval_date, str):
            try:
                eval_date = datetime.fromisoformat(eval_date.replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except:
                pass
    
    resume_data = [
        ("Fecha", eval_date or "N/A"),
        ("Porcentaje General", f"{general_pct:.1f}%"),
        ("Porcentaje PA (Gestión)", f"{pa_pct:.1f}%"),
        ("Porcentaje PO (Operaciones)", f"{po_pct:.1f}%"),
        ("Prioridad", priority),
    ]
    
    for idx, (label, value) in enumerate(resume_data, start=3):
        _create_styled_cell(ws1, idx, 1, label, bold=True)
        ws1.cell(row=idx, column=2, value=value)
    
    # Set column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    
    # ========== Sheet 2: Desglose PA ==========
    ws2 = wb.create_sheet("Desglose PA")
    
    ws2.merge_cells('A1:C1')
    header_cell = ws2.cell(row=1, column=1, value="DESGLOSE PA - GESTIÓN")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    _style_header(ws2, 2, 3)
    ws2.cell(row=2, column=1, value="Aspecto")
    ws2.cell(row=2, column=2, value="Porcentaje")
    ws2.cell(row=2, column=3, value="Estado")
    
    # Data
    row = 3
    for aspect, pct in pa_breakdown.items():
        ws2.cell(row=row, column=1, value=aspect)
        ws2.cell(row=row, column=2, value=f"{pct:.1f}%")
        ws2.cell(row=row, column=3, value=_get_score_status(pct))
        row += 1
    
    # Set column widths
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 18
    
    # ========== Sheet 3: Desglose PO ==========
    ws3 = wb.create_sheet("Desglose PO")
    
    ws3.merge_cells('A1:C1')
    header_cell = ws3.cell(row=1, column=1, value="DESGLOSE PO - OPERACIONES")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    _style_header(ws3, 2, 3)
    ws3.cell(row=2, column=1, value="Aspecto")
    ws3.cell(row=2, column=2, value="Porcentaje")
    ws3.cell(row=2, column=3, value="Estado")
    
    # Data
    row = 3
    for aspect, pct in po_breakdown.items():
        ws3.cell(row=row, column=1, value=aspect)
        ws3.cell(row=row, column=2, value=f"{pct:.1f}%")
        ws3.cell(row=row, column=3, value=_get_score_status(pct))
        row += 1
    
    # Set column widths
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 18
    
    # ========== Sheet 4: Detalle de Preguntas ==========
    ws4 = wb.create_sheet("Detalle de Preguntas")
    
    ws4.merge_cells('A1:E1')
    header_cell = ws4.cell(row=1, column=1, value="DETALLE DE PREGUNTAS")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    _style_header(ws4, 2, 5)
    ws4.cell(row=2, column=1, value="Aspecto")
    ws4.cell(row=2, column=2, value="Pregunta")
    ws4.cell(row=2, column=3, value="Contexto")
    ws4.cell(row=2, column=4, value="Puntuación")
    ws4.cell(row=2, column=5, value="Porcentaje")
    
    # Data
    row = 3
    for q in questions:
        ws4.cell(row=row, column=1, value=q["aspect"])
        ws4.cell(row=row, column=2, value=q["question"])
        ws4.cell(row=row, column=3, value=q.get("context", ""))
        ws4.cell(row=row, column=4, value=f"{q['rating']}/5")
        ws4.cell(row=row, column=5, value=f"{q['percentage']:.1f}%")
        row += 1
    
    # Set column widths
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 50
    ws4.column_dimensions['C'].width = 30
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 12
    
    # ========== Sheet 5: Recomendaciones ==========
    ws5 = wb.create_sheet("Recomendaciones")
    
    ws5.merge_cells('A1:C1')
    header_cell = ws5.cell(row=1, column=1, value="RECOMENDACIONES")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    _style_header(ws5, 2, 3)
    ws5.cell(row=2, column=1, value="Prioridad")
    ws5.cell(row=2, column=2, value="Elemento")
    ws5.cell(row=2, column=3, value="Recomendación")
    
    # Data
    row = 3
    for rec in recommendations:
        ws5.cell(row=row, column=1, value=rec["priority"])
        ws5.cell(row=row, column=2, value=rec["element"])
        ws5.cell(row=row, column=3, value=rec["recommendation"])
        row += 1
    
    # Set column widths
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 35
    ws5.column_dimensions['C'].width = 80
    
    # ========== Sheet 6: Planes de Acción ==========
    ws6 = wb.create_sheet("Planes de Acción")
    
    ws6.merge_cells('A1:E1')
    header_cell = ws6.cell(row=1, column=1, value="PLANES DE ACCIÓN")
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    _style_header(ws6, 2, 5)
    ws6.cell(row=2, column=1, value="Acción")
    ws6.cell(row=2, column=2, value="Estado")
    ws6.cell(row=2, column=3, value="Responsable")
    ws6.cell(row=2, column=4, value="Fecha Límite")
    ws6.cell(row=2, column=5, value="Descripción")
    
    # Data
    row = 3
    if action_plans:
        for plan in action_plans:
            ws6.cell(row=row, column=1, value=plan.get("action", ""))
            ws6.cell(row=row, column=2, value=plan.get("status", ""))
            ws6.cell(row=row, column=3, value=plan.get("responsible", ""))
            ws6.cell(row=row, column=4, value=plan.get("due_date", ""))
            ws6.cell(row=row, column=5, value=plan.get("description", ""))
            row += 1
    else:
        ws6.cell(row=3, column=1, value="No hay planes de acción registrados")
        ws6.merge_cells('A3:E3')
    
    # Set column widths
    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 15
    ws6.column_dimensions['C'].width = 20
    ws6.column_dimensions['D'].width = 15
    ws6.column_dimensions['E'].width = 50
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=auditoria-{evaluation_id}.xlsx"}
    )